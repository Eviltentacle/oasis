import argparse
import base64
import logging
import shutil
import signal
import sys
import time
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
import pandas as pd
import requests
from dotenv import load_dotenv

load_dotenv()
API_KEY = os.getenv("API_KEY")
if not API_KEY:
    raise ValueError("VirusTotal API key not found. Please set API_KEY in a .env file.")

INPUT_FILE = 'ioc_vetting.xlsx'
processed_hashes = set()

# Handle Ctrl+C
def sigint_handler(signal_received, frame):
    print("\n[!] Script interrupted by user. Processed hashes saved.")
    sys.exit(0)

signal.signal(signal.SIGINT, sigint_handler)

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def check_hash(hash_value):
    url = f"https://www.virustotal.com/api/v3/files/{hash_value}"
    headers = {
        "accept": "application/json",
        "x-apikey": API_KEY
    }

    try:
        response = requests.get(url, headers=headers, verify=False)

        if response.status_code == 404:
            # Hash not found in VT
            return [hash_value, 'Not Found', '', '', '', '']

        response.raise_for_status()
        attributes = response.json()['data']['attributes']

        filename = attributes.get('meaningful_name', '')
        sha256 = attributes.get('sha256', '')
        sha1 = attributes.get('sha1', '')
        stats = attributes.get('last_analysis_stats', {})
        malicious = stats.get('malicious', 0)
        suspicious = stats.get('suspicious', 0)
        total_malicious = int(malicious) + int(suspicious)

        analysis = attributes.get('last_analysis_results', {})
        sentinelone = analysis.get('SentinelOne', {}).get('category', '')

        return [
            filename,  # A
            sha1,      # B
            sha256,    # C
            total_malicious,  # D
            sentinelone,      # E
        ]

    except Exception as e:
        logging.error(f"Failed to process hash {hash_value}: {e}")
        # Return error in place of all fields except original hash
        return [hash_value, 'Error', '', '', '', '']

def main():
    parser = argparse.ArgumentParser(description="VirusTotal Hash Scanner (Excel-based)")
    #parser.add_argument('-i', '--input-file', required=True, help='Input Excel template file')
    parser.add_argument('-o', '--output-file', required=True, help='Output Excel file (copy of input)')
    args = parser.parse_args()

    if not os.path.exists(args.output_file):
        shutil.copyfile(INPUT_FILE, args.output_file)
        logging.info(f"Copied template '{INPUT_FILE}' to '{args.output_file}'")
    else:
        logging.info(f"Output file '{args.output_file}' already exists. Will update 'Hash' sheet only.")

    try:
        start_time = time.time()
        df_hashes = pd.read_excel(args.output_file, sheet_name='Hash')
        df_hashes.columns = df_hashes.columns.str.strip()

        if 'Name' not in df_hashes.columns:
            logging.error("Missing 'Name' column in 'Hash' sheet.")
            return

        hash_list = df_hashes['Name'].dropna().astype(str).tolist()
        logging.info(f"Loaded {len(hash_list)} hashes.")

        unique_hashes = list(dict.fromkeys(hash_list))

        result_map = {}

        with ThreadPoolExecutor(max_workers=10) as executor:
            future_to_hash = {executor.submit(check_hash, h): h for h in unique_hashes}
            for future in as_completed(future_to_hash):
                h = future_to_hash[future]
                try:
                    result_map[h] = future.result()
                    #logging.info(f"Processed hash: {hash_list}")
                except Exception as e:
                    logging.error(f"Error processing hash {h}: {e}")
                    result_map[h] = [h, 'Error', '', '', '', '']

        workbook = load_workbook(args.output_file)
        sheet = workbook['Hash']

        start_row = 2

        for idx, hash_val in enumerate(hash_list):
            row_num = start_row + idx
            res = result_map.get(hash_val, [hash_val, 'Not Found', '', '', '', ''])
            for col_idx, value in enumerate(res, start=1):
                sheet.cell(row=row_num, column=col_idx, value=value)

        workbook.save(args.output_file)
        logging.info(f"Results written to 'Hash' sheet in '{args.output_file}'")
        print(f"\nScan completed in {round(time.time() - start_time, 2)} seconds.")
    except Exception as e:
        logging.error(f"Unexpected error: {e}")

if __name__ == '__main__':
    main()



