import argparse
import pandas as pd
import requests
import base64
import time
import shutil
import logging
import os
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook

API_KEY = 'YOUR_API'  # <-- Replace with your VirusTotal Premium API key
INPUT_FILE = 'ioc_vetting.xlsx'

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def check_input(value):
    vt_base_url = "https://www.virustotal.com/api/v3"
    headers = {'x-apikey': API_KEY}
    is_url = value.startswith("http://") or value.startswith("https://")

    try:
        if is_url:
            url_id = base64.urlsafe_b64encode(value.encode()).decode().strip("=")
            endpoint = f"{vt_base_url}/urls/{url_id}"
            response = requests.get(endpoint, headers=headers, verify=False)
            if response.status_code == 404:
                return [value, "Not found", ""]
            response.raise_for_status()
            stats = response.json().get("data", {}).get("attributes", {}).get("last_analysis_stats", {})
            malicious = stats.get("malicious", 0)
        else:
            parsed = urlparse("http://" + value if "://" not in value else value)
            url = parsed.hostname or value
            endpoint = f"{vt_base_url}/urls/{url}"
            response = requests.get(endpoint, headers=headers)
            if response.status_code == 404:
                return [url, "Not found", ""]
            response.raise_for_status()
            stats = response.json().get("data", {}).get("attributes", {}).get("last_analysis_stats", {})
            malicious = stats.get("malicious", 0)

        return [value, malicious, ""]

    except requests.exceptions.RequestException as e:
        logging.error(f"Request failed for {value}: {e}")
        return [value, "Request error", ""]

def main():
    parser = argparse.ArgumentParser(description='VirusTotal url/URL Scanner (Excel-based)')
    #parser.add_argument('-i', '--input-file', required=True, help='Input Excel template file')
    parser.add_argument('-o', '--output-file', required=True, help='Output Excel file (copy of input)')
    args = parser.parse_args()

    if not os.path.exists(args.output_file):
        shutil.copyfile(INPUT_FILE, args.output_file)
        logging.info(f"Copied template '{INPUT_FILE}' to '{args.output_file}'")
    else:
        logging.info(f"Output file '{args.output_file}' already exists. Will update 'url' sheet only.")

    try:
        start_time = time.time()
        df_url = pd.read_excel(args.output_file, sheet_name='URL')
        df_url.columns = df_url.columns.str.strip()

        if 'URL' not in df_url.columns:
            logging.error("Missing 'url' column in 'url' sheet.")
            return

        urls = df_url['URL'].dropna().astype(str).tolist()
        logging.info(f"Loaded {len(urls)} urls/URLs from url sheet.")

        results = []
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = {executor.submit(check_input, url): url for url in urls}
            for future in as_completed(futures):
                results.append(future.result())
                #logging.info(f"Processed URL: {urls}")

        workbook = load_workbook(args.output_file)
        sheet = workbook['URL']

        start_row = 2 
        for i, result in enumerate(results, start=start_row):
            for j, value in enumerate(result, start=1): 
                sheet.cell(row=i, column=j, value=value)

        workbook.save(args.output_file)
        logging.info(f"Scanned Completed. Results saved to '{args.output_file}'")
        print(f"\nScan completed in {round(time.time() - start_time, 2)} seconds.")
    except Exception as e:
        logging.error(f"Unexpected error: {e}")

if __name__ == '__main__':
    main()



